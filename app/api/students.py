import csv
from io import BytesIO, StringIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, false, or_

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import asc
from sqlalchemy.orm import Session, joinedload

from app.api.deps import get_current_user
from app.db.deps import get_db
from app.models.parent_student import ParentStudent
from app.models.student import Student
from app.models.teacher_assignment import TeacherAssignment
from app.models.user import User
from app.schemas.auth import UserOut
from app.schemas.classroom import normalize_class_letter, validate_grade_range
from app.schemas.student import (
    ClassOptionsResponse,
    ParentStudentCreate,
    ParentStudentOut,
    StudentCreate,
    StudentListResponse,
    StudentOut,
    StudentUpdate,
)

router = APIRouter(prefix="/student", tags=["Students"])


class NotificationRequests(BaseModel):
    subject: str
    message: str


def require_admin(current_user: User) -> None:
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")


def build_class_suffix(grade: int | None, class_letter: str | None) -> str:
    if not grade or not class_letter:
        return ""
    class_letter_map = str.maketrans(
        {
            "А": "A",
            "В": "B",
            "Е": "E",
            "К": "K",
            "М": "M",
            "Н": "N",
            "О": "O",
            "Р": "P",
            "С": "S",
            "Т": "T",
            "У": "Y",
            "Х": "X",
        }
    )
    letter = class_letter.strip().upper().translate(class_letter_map)
    return f"_{grade}{letter}"


def parent_can_access_student(db: Session, parent: User, student_id: str) -> bool:
    return (
        db.query(ParentStudent.id)
        .join(Student, Student.id == ParentStudent.student_id)
        .filter(
            ParentStudent.parent_id == parent.id,
            ParentStudent.student_id == student_id,
            Student.school_id == parent.school_id,
        )
        .first()
        is not None
    )


def teacher_class_clauses(db: Session, teacher: User):
    assignments = (
        db.query(TeacherAssignment.grade, TeacherAssignment.class_letter)
        .filter(
            TeacherAssignment.teacher_id == teacher.id,
            TeacherAssignment.school_id == teacher.school_id,
        )
        .distinct()
        .all()
    )
    if not assignments:
        return [false()]
    return [
        and_(Student.grade == grade, Student.class_letter == class_letter)
        for grade, class_letter in assignments
    ]


def teacher_can_access_student(db: Session, teacher: User, student: Student) -> bool:
    return (
        db.query(TeacherAssignment.id)
        .filter(
            TeacherAssignment.teacher_id == teacher.id,
            TeacherAssignment.school_id == teacher.school_id,
            TeacherAssignment.grade == student.grade,
            TeacherAssignment.class_letter == student.class_letter,
        )
        .first()
        is not None
    )


def auto_link_parent_by_email(db: Session, student: Student) -> None:
    parent = (
        db.query(User)
        .filter(
            User.school_id == student.school_id,
            User.role == "parent",
            User.email.ilike(student.email),
        )
        .first()
    )
    if not parent:
        return

    existing = (
        db.query(ParentStudent)
        .filter(
            ParentStudent.parent_id == parent.id,
            ParentStudent.student_id == student.id,
        )
        .first()
    )
    if existing:
        return

    db.add(
        ParentStudent(
            parent_id=parent.id,
            student_id=student.id,
            relationship="Родитель",
        )
    )


@router.get("/", response_model=StudentListResponse)
def get_students(
    grade: Optional[int] = None,
    class_letter: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")

    query = db.query(Student).filter(Student.school_id == current_user.school_id)
    normalized_class_letter = None
    if class_letter is not None:
        try:
            normalized_class_letter = normalize_class_letter(class_letter)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid class letter") from None

    if current_user.role == "parent":
        query = query.join(ParentStudent).filter(ParentStudent.parent_id == current_user.id)
    elif current_user.role == "teacher":
        query = query.filter(or_(*teacher_class_clauses(db, current_user)))

    if grade is not None:
        try:
            grade = validate_grade_range(grade)
        except ValueError:
            raise HTTPException(status_code=400, detail="Grade must be between 1 and 11") from None
        query = query.filter(Student.grade == grade)

    if normalized_class_letter is not None:
        query = query.filter(Student.class_letter == normalized_class_letter)

    if search and search.strip():
        search_terms = search.strip().split()
        query = query.filter(
            and_(
                *[
                    or_(
                        Student.first_name.ilike(f"{term}%"),
                        Student.last_name.ilike(f"{term}%"),
                        Student.middle_name.ilike(f"{term}%"),
                        Student.email.ilike(f"{term}%"),
                    )
                    for term in search_terms
                ]
            )
        )

    total = query.count()
    offset = (page - 1) * limit
    pages = (total + limit - 1) // limit if total > 0 else 0

    items = (
        query.order_by(asc(Student.last_name), asc(Student.first_name))
        .offset(offset)
        .limit(limit)
        .all()
    )

    return {
        "items": items,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": pages,
    }


@router.get("/class-options", response_model=ClassOptionsResponse)
def get_class_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")

    query = (
        db.query(Student.grade, Student.class_letter)
        .filter(Student.school_id == current_user.school_id)
        .distinct()
    )

    if current_user.role == "parent":
        query = query.join(ParentStudent).filter(ParentStudent.parent_id == current_user.id)
    elif current_user.role == "teacher":
        query = query.filter(or_(*teacher_class_clauses(db, current_user)))

    rows = query.order_by(asc(Student.grade), asc(Student.class_letter)).all()
    classes = [
        {"grade": grade, "class_letter": class_letter}
        for grade, class_letter in rows
    ]

    return {
        "grades": sorted({grade for grade, _ in rows}),
        "letters": sorted({class_letter for _, class_letter in rows}),
        "classes": classes,
    }


@router.get("/export")
def export_students(
    grade: Optional[int] = None,
    class_letter: Optional[str] = None,
    format: str = Query(default="xlsx", pattern="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")
    require_admin(current_user)

    query = db.query(Student).filter(Student.school_id == current_user.school_id)
    if grade is not None:
        try:
            grade = validate_grade_range(grade)
        except ValueError:
            raise HTTPException(status_code=400, detail="Grade must be between 1 and 11") from None
        query = query.filter(Student.grade == grade)
    if class_letter is not None:
        try:
            class_letter = normalize_class_letter(class_letter)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid class letter") from None
        query = query.filter(Student.class_letter == class_letter)

    students = query.order_by(asc(Student.grade), asc(Student.class_letter), asc(Student.last_name)).all()
    filename_class = build_class_suffix(grade, class_letter)

    headers = ["last_name", "first_name", "middle_name", "email", "grade", "class_letter"]
    rows = [
        [
            student.last_name,
            student.first_name,
            student.middle_name,
            student.email,
            student.grade,
            student.class_letter,
        ]
        for student in students
    ]

    if format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return StreamingResponse(
            iter([output.getvalue().encode("utf-8-sig")]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="students{filename_class}.csv"'
            },
        )

    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ученики"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for column, width in {"A": 20, "B": 18, "C": 22, "D": 30, "E": 10, "F": 14}.items():
        sheet.column_dimensions[column].width = width
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="students{filename_class}.xlsx"'
        },
    )


@router.post("/import")
def import_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")
    require_admin(current_user)

    raw = file.file.read()
    name = (file.filename or "").lower()
    rows: list[dict[str, str]] = []

    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(text)))
    elif name.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if values:
            headers = [str(value).strip() for value in values[0]]
            for row in values[1:]:
                rows.append(
                    {
                        headers[index]: "" if value is None else str(value).strip()
                        for index, value in enumerate(row)
                        if index < len(headers)
                    }
                )
    else:
        raise HTTPException(status_code=400, detail="Only CSV and XLSX files are supported")

    required = {"last_name", "first_name", "middle_name", "email", "grade", "class_letter"}
    created = 0
    skipped = 0
    errors: list[str] = []

    for index, row in enumerate(rows, start=2):
        normalized = {str(key).strip(): value for key, value in row.items() if key is not None}
        if not required.issubset(normalized.keys()):
            skipped += 1
            errors.append(f"Row {index}: missing required columns")
            continue

        try:
            grade = validate_grade_range(int(str(normalized["grade"]).strip()))
            class_letter = normalize_class_letter(str(normalized["class_letter"]))
        except ValueError:
            skipped += 1
            errors.append(f"Row {index}: invalid grade or class_letter")
            continue

        student = Student(
            first_name=str(normalized["first_name"]).strip(),
            last_name=str(normalized["last_name"]).strip(),
            middle_name=str(normalized["middle_name"]).strip(),
            email=str(normalized["email"]).strip(),
            grade=grade,
            class_letter=class_letter,
            school_id=current_user.school_id,
        )
        db.add(student)
        db.flush()
        auto_link_parent_by_email(db, student)
        created += 1

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors[:20]}


@router.get("/{student_id}", response_model=StudentOut)
def get_student_by_id(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")

    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.school_id == current_user.school_id)
        .first()
    )

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if current_user.role == "parent" and not parent_can_access_student(db, current_user, student.id):
        raise HTTPException(status_code=404, detail="Student not found")
    if current_user.role == "teacher" and not teacher_can_access_student(db, current_user, student):
        raise HTTPException(status_code=404, detail="Student not found")

    return student


@router.get("/{student_id}/parents", response_model=list[ParentStudentOut])
def get_student_parents(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")

    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.school_id == current_user.school_id)
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if current_user.role == "parent" and not parent_can_access_student(db, current_user, student.id):
        raise HTTPException(status_code=404, detail="Student not found")
    if current_user.role == "teacher" and not teacher_can_access_student(db, current_user, student):
        raise HTTPException(status_code=404, detail="Student not found")

    return (
        db.query(ParentStudent)
        .options(joinedload(ParentStudent.parent))
        .filter(ParentStudent.student_id == student.id)
        .join(User, User.id == ParentStudent.parent_id)
        .order_by(asc(User.last_name), asc(User.first_name))
        .all()
    )


@router.get("/{student_id}/parents/available", response_model=list[UserOut])
def get_available_parents(
    student_id: str,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")
    require_admin(current_user)

    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.school_id == current_user.school_id)
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    linked_parent_ids = (
        db.query(ParentStudent.parent_id)
        .filter(ParentStudent.student_id == student.id)
        .subquery()
    )
    query = (
        db.query(User)
        .filter(
            User.school_id == current_user.school_id,
            User.role == "parent",
            User.id.notin_(linked_parent_ids),
        )
    )
    if search and search.strip():
        search_terms = search.strip().split()
        query = query.filter(
            and_(
                *[
                    or_(
                        User.first_name.ilike(f"{term}%"),
                        User.last_name.ilike(f"{term}%"),
                        User.middle_name.ilike(f"{term}%"),
                        User.email.ilike(f"{term}%"),
                    )
                    for term in search_terms
                ]
            )
        )

    return query.order_by(asc(User.last_name), asc(User.first_name)).limit(20).all()


@router.post("/{student_id}/parents", response_model=ParentStudentOut, status_code=201)
def attach_parent_to_student(
    student_id: str,
    data: ParentStudentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")
    require_admin(current_user)

    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.school_id == current_user.school_id)
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    parent = (
        db.query(User)
        .filter(
            User.id == data.parent_id,
            User.school_id == current_user.school_id,
            User.role == "parent",
        )
        .first()
    )
    if not parent:
        raise HTTPException(status_code=404, detail="Parent not found")

    existing = (
        db.query(ParentStudent)
        .filter(
            ParentStudent.parent_id == parent.id,
            ParentStudent.student_id == student.id,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Parent is already linked")

    link = ParentStudent(
        parent_id=parent.id,
        student_id=student.id,
        relationship=data.relationship.strip() if data.relationship else None,
    )
    db.add(link)
    db.commit()
    db.refresh(link)
    return (
        db.query(ParentStudent)
        .options(joinedload(ParentStudent.parent))
        .filter(ParentStudent.id == link.id)
        .first()
    )


@router.delete("/{student_id}/parents/{parent_id}", status_code=204)
def detach_parent_from_student(
    student_id: str,
    parent_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")
    require_admin(current_user)

    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.school_id == current_user.school_id)
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    link = (
        db.query(ParentStudent)
        .join(User, User.id == ParentStudent.parent_id)
        .filter(
            ParentStudent.student_id == student.id,
            ParentStudent.parent_id == parent_id,
            User.school_id == current_user.school_id,
        )
        .first()
    )
    if not link:
        raise HTTPException(status_code=404, detail="Parent link not found")

    db.delete(link)
    db.commit()


@router.patch("/{student_id}", response_model=StudentOut)
def update_student(
    student_id: str,
    data: StudentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")
    require_admin(current_user)

    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.school_id == current_user.school_id)
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    if data.first_name is not None:
        student.first_name = data.first_name
    if data.last_name is not None:
        student.last_name = data.last_name
    if data.middle_name is not None:
        student.middle_name = data.middle_name
    if data.email is not None:
        student.email = data.email
    if data.grade is not None:
        student.grade = data.grade
    if data.class_letter is not None:
        student.class_letter = data.class_letter

    db.commit()
    db.refresh(student)
    auto_link_parent_by_email(db, student)
    db.commit()
    return student


@router.post("/", response_model=StudentOut)
def create_students(
    data: StudentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")
    require_admin(current_user)

    student = Student(
        first_name=data.first_name,
        last_name=data.last_name,
        middle_name=data.middle_name,
        email=data.email,
        grade=data.grade,
        class_letter=data.class_letter,
        school_id=current_user.school_id,
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    auto_link_parent_by_email(db, student)
    db.commit()
    return student


@router.delete("/{student_id}", status_code=204)
def delete_student(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.school_id:
        raise HTTPException(status_code=403, detail="User is not linked to a school")
    require_admin(current_user)

    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.school_id == current_user.school_id)
        .first()
    )

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    db.delete(student)
    db.commit()
